# -*- coding: UTF-8 -*-
import json
import time
import requests
import math
import pandas as pd


# import xlrd
from openpyxl import load_workbook, Workbook
# from xlutils.copy import copy
# import os
# import sys
# os.environ['REQUESTS_CA_BUNDLE'] = os.path.join(
#     os.path.dirname(sys.argv[0]), 'cacert.pem')
# importlib.reload(sys)
# sys.setdefaultencoding('utf8')
# import certifi
# print(certifi.where())


s = requests.Session()


def get_reviews(appid, params):
    url = 'https://store.steampowered.com/appreviews/'
    requests.adapters.DEFAULT_RETRIES = 5  # 增加重连次数
    s.keep_alive = False
    try:
        response = s.get(url=url+appid, params=params,
                         headers={'User-Agent': 'Mozilla/5.0', 'Connection': 'close'})
    except Exception as err:
        print(err, '-=========err')
        pass

    # print(response, '-=========response')
    return response.json()

# params = {'json': 1}
# response = get_reviews(775500, params)
# cursor = response['cursor']
# params['cursor'] = cursor.encode()
# response_2 = get_reviews(775500, params)


# def get_n_reviews(appid, n=100):
def get_n_reviews(appid, n):

    wb = Workbook()
    # 在索引为0的位置创建一个名为steamReviews的sheet页
    ws = wb.create_sheet('steamReviews_' + str(appid), 0)
    # 对sheet页设置一个颜色（16位的RGB颜色）
    ws.sheet_properties.tabColor = 'ff72BA'
    # 将创建的工作簿保存为steamReviews.xlsx
    wb.save('steamReviews_' + str(appid)+'.xlsx')
    # 最后关闭文件
    wb.close()
    reviews = []
    query_summary = ''
    total_reviews = 0
    num = 0
    cursor = '*'
    index = 0
    params = {
        'json': 1,
        'filter': 'updated',
        'language': 'all',
        # 'day_range': 9223372036854775807,
        'review_type': 'all',
        'purchase_type': 'all'
    }

    while n > 0:
        params['cursor'] = cursor.encode()
        params['num_per_page'] = min(100, n)

        response = get_reviews(appid, params)
        # print(response, 'response=================')

        # if response['query_summary']:
        #     query_summary = response['query_summary']

        # if query_summary:
        #     total_reviews = query_summary['total_reviews']
        #     n = query_summary['total_reviews'] - 100
        # else:
        #     n -= 100
        if 'query_summary' in response:
            query_summary = response['query_summary']
            if 'total_reviews' in query_summary:
                total_reviews = query_summary['total_reviews']
            print(total_reviews, '剩余多少评测，total_reviews=================')
        if total_reviews > 100:
            total_reviews -= 100
            n = total_reviews
            print(n//100, '第n个100条==========config')
        elif total_reviews < 100 and total_reviews > 0:
            # total_reviews -= 100
            n = total_reviews
            print(n//100, '第n个100条==========小于100')
        else:
            break
        if total_reviews == 0:
            n -= 100
        cursor = response['cursor']
        reviews += response['reviews']

        index = save_data(response, index, appid)
        time.sleep(1)
        if len(response['reviews']) < 100:
            break

    # return reviews, query_summary
    # print(reviews, 'reviews=================')


def save_data(reviews, index, appid):
    res = reviews
    tempQuery_summary = ""
    # index = len(reviews)

    # print(len(reviews['reviews']), '============index = len(reviews)')
    list = []
    sum_list = []
    for d in res["reviews"]:

        # print(d, '============d')
        project = {}
        project["recommendationid"] = d["recommendationid"]
        project["language"] = d["language"]

        project["author.steamid"] = d["author"]["steamid"]
        project["author.num_games_owned"] = d["author"]["num_games_owned"]
        project["author.num_reviews"] = d["author"]["num_reviews"]
        project["author.playtime_forever"] = d["author"]["playtime_forever"]
        project["author.playtime_last_two_weeks"] = d["author"]["playtime_last_two_weeks"]
        project["author.playtime_at_review"] = d["author"]["playtime_at_review"]
        project["author.last_played"] = time.strftime(
            "%Y-%m-%d %H:%M:%S", time.localtime(d["author"]["last_played"]))

        project["voted_up"] = '推荐'if d["voted_up"] else '不推荐'
        project["votes_up"] = d["votes_up"]
        project["votes_funny"] = d["votes_funny"]
        project["steam_purchase"] = d["steam_purchase"]
        project["review"] = d["review"]
        list.append(project)
    # 导出Excel格式的文件
    if 'query_summary' in res:
        tempQuery_summary = res['query_summary']
        if 'total_reviews' in tempQuery_summary:
            tempproject = {}
            # tempQuery_summary
            # project["num_reviews"] = tempQuery_summary["num_reviews"]
            tempproject["review_score"] = tempQuery_summary["review_score"]
            tempproject["review_score_desc"] = tempQuery_summary["review_score_desc"]
            tempproject["total_positive"] = tempQuery_summary["total_positive"]
            tempproject["total_negative"] = tempQuery_summary["total_negative"]
            tempproject["total_reviews"] = tempQuery_summary["total_reviews"]
            sum_list.append(tempproject)
    # 方法一：
    # 解决乱码问题
    # print(json.dumps(list, ensure_ascii=False))
    # 'steamReviews_' + str(appid)+'.xlsx'
    # wb = load_workbook(r"F:\democode\steam\test.xlsx")
    wb = load_workbook(r"steamReviews_" + str(appid)+".xlsx")
    # f = open(r"response1.json", "a", encoding="UTF-8")
    f = open(r"response_"+str(appid)+".json", "a", encoding="UTF-8")
    json.dump(list, f, ensure_ascii=False, indent=3)
    f.close()
    # F:\democode\steam\steamApi.py
    # sheet = wb["Sheet1"]
    sheet = wb['steamReviews_' + str(appid)]

    start_row_num = 1
    # index = 0
    sheet.cell(1, 1).value = "该推荐的唯一 ID --- recommendationid "
    sheet.cell(1, 2).value = "用户撰写评测时使用的语言 --- language "

    sheet.cell(1, 3).value = "用户的 SteamID --- steamid "
    sheet.cell(1, 4).value = "用户拥有的游戏数量 --- num_games_owned "
    sheet.cell(1, 5).value = "用户撰写的评测数量 --- num_reviews "
    sheet.cell(1, 6).value = "此应用所记录的总计游戏时间（分钟） --- playtime_forever "
    sheet.cell(1, 7).value = "此应用所记录的过去两周的游戏时间（分钟） --- playtime_last_two_weeks "
    sheet.cell(1, 8).value = "撰写评测时的游戏时间（分钟） --- playtime_at_review "
    sheet.cell(1, 9).value = "用户上次游戏的时间（日期） --- last_played "

    sheet.cell(1, 10).value = "true 表示是一则正面推荐 --- voted_up "
    sheet.cell(1, 11).value = "认为这篇评测有价值的用户人数 --- votes_up "
    sheet.cell(1, 12).value = "认为这篇评测欢乐的用户人数 --- votes_funny "
    sheet.cell(1, 13).value = "true 表示用户在 Steam 上购买了此游戏 --- steam_purchase "
    sheet.cell(1, 14).value = "评测文本 --- review "
    sheet.cell(1, 15).value = "游戏总体的 评测分数 --- review_score"
    sheet.cell(1, 16).value = "游戏总体的 评测分数描述 --- review_score_desc"
    sheet.cell(1, 17).value = "游戏总体的 正面评测的总数 --- total_positive"
    sheet.cell(1, 18).value = "游戏总体的 负面评测的总数 --- total_negative"
    sheet.cell(1, 19).value = "游戏总体的 符合查询参数的评测总数 --- total_reviews"
    # sheet.cell(1, 19).value = "游戏总体的 --- total_reviews"
    # sheet.cell(1, 20).value = "评测文本 --- "
    # 把数值填充到excel表格中
    for one in list:
        index += 1
        column_index = 1
        for key, value in one.items():
            sheet.cell(start_row_num + index, column_index).value = value
            column_index = column_index + 1
    if len(sum_list) > 0:
        for one in sum_list:
            column_index = 15
            for key, value in one.items():
                sheet.cell(2, column_index).value = value
                column_index = column_index + 1
    # wb.save(r"F:\democode\steam\test.xlsx")
    wb.save(r"steamReviews_" + str(appid)+".xlsx")
    # print(index, '=============index')
    print("文件保存成功！")
    return index


def read_config(index):
    """"读取配置"""
    with open("output19.json", encoding="UTF-8") as json_file:
        config = json.load(json_file)
        # print(config, '==========config')
        print(save_data(config, index)+1)


def demo(ddd):
    n = 100
    print(n, '1111==========n')
    total_reviews = 0
    while n > 0:
        if 'query_summary' in ddd:
            query_summary = ddd['query_summary']
            total_reviews = query_summary['total_reviews']
        if total_reviews > 100:
            total_reviews -= 100
            n = total_reviews
            print(n, '222==========config')
        else:
            break
        if total_reviews == 0:
            n -= 100
        ddd = {
            "name": "Bill",
            "age": 63,
            "city": "Seatle"
        }
        print(n, '33333==========config')


# demo({
#     "query_summary": {'num_reviews': 100, 'review_score': 8, 'review_score_desc': 'Very Positive', 'total_positive': 8601, 'total_negative': 1502, 'total_reviews': 10103},
# })
# read_config(0)
# read_config(0)
# get_n_reviews("775500", 100)

def startInput():
    print('输入需要查询的steam游戏AppId，按Q退出\n')
    while True:
        n = input()
        if n == 'Q':
            break
        elif n:
            print('输入需要查询的steam游戏AppId', n)
            get_n_reviews(str(n), 100)
            break
        else:
            print('输入错误')


if __name__ == "__main__":
    startInput()

# def ddd():
#     num = 1
#     print('steamReviews' + str(num)+'html')


# ddd()
